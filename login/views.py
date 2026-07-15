import json
import decimal
import random
import urllib.error
import urllib.request
from datetime import date, timedelta
from decimal import Decimal
from math import ceil, pi

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.password_validation import validate_password
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.core.mail import EmailMultiAlternatives
from django.core.validators import validate_email
from django.db.models import Sum, Q
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.template.loader import render_to_string
from django.utils import timezone
from django.utils.html import strip_tags
from django.views.decorators.http import require_GET, require_POST, require_http_methods

from .models import Transaction, UserProfile, SavingsGoal
from .spending_coach import fetch_dad_joke, fetch_motivation, get_spending_message
from .templatetags.currency_filters import indian_currency


# ── Smart Savings Motivation ──────────────────────────────
_MOTIVATION_QUOTES = {
    "broke": [
        (
            "₹0 saved? Bold strategy. Let's see if it pays off.",
            "— Your Future Self, Nervous",
        ),
        (
            "Your savings account called. It said it's lonely and considering therapy.",
            "— Bank of Feelings",
        ),
        (
            "Not saving is just borrowing stress from your future self. He's not happy.",
            "— Future You, Crying",
        ),
        (
            "The good news: you can only go up from here. The bad news: you're here.",
            "— Motivational Rock Bottom",
        ),
        (
            "Your wallet is on a juice cleanse. Involuntarily.",
            "— Financial Wellness Guru",
        ),
    ],
    "just_started": [
        (
            "You saved something! It's not much, but neither was the first step on the moon.",
            "— Neil Armstrong, Probably",
        ),
        (
            "A tiny saving today is a slightly less tiny saving tomorrow. Math!",
            "— Einstein's Finance Cousin",
        ),
        (
            "You've started! That's more than 90% of people who said 'I'll start Monday'.",
            "— The Monday Club",
        ),
        (
            "Your savings are like a baby — small, fragile, but full of potential. Don't drop it.",
            "— Parenting & Finance Weekly",
        ),
        (
            "Progress: tiny. Attitude: massive. Keep going, legend.",
            "— Your Hype Person",
        ),
    ],
    "making_progress": [
        (
            "You're not broke, you're 'pre-wealthy'. Keep going.",
            "— Optimism Department",
        ),
        (
            "Halfway there! Your future self just sent a thumbs up emoji.",
            "— Time Travel Update",
        ),
        (
            "You're saving money like it owes you something. It does. Keep at it.",
            "— Debt Collector (Nice Version)",
        ),
        (
            "Look at you, being financially responsible. Your parents would be shocked.",
            "— Family Group Chat",
        ),
        (
            "You're in the top 40% of people who actually save. The bar is low, but you cleared it!",
            "— Statistics That Slap",
        ),
    ],
    "almost_there": [
        (
            "So close! Don't you dare buy that thing you're thinking about right now.",
            "— Your Conscience",
        ),
        (
            "You're 80%+ there. This is not the time to 'treat yourself'.",
            "— The Voice of Reason",
        ),
        (
            "Almost at your goal. One more month of pretending you don't need takeout.",
            "— Delivery App, Sad",
        ),
        (
            "The finish line is RIGHT THERE. Don't look at the sale section.",
            "— Coach Willpower",
        ),
        (
            "You've come too far to spend it on something you'll return in 3 days.",
            "— Return Policy Dept.",
        ),
    ],
    "goal_reached": [
        (
            "GOAL REACHED! You did it! Now set a harder one so we can do this again.",
            "— Masochism & Finance",
        ),
        (
            "You actually saved the money. We're as surprised as you are. Congrats!",
            "— SpendWise HQ",
        ),
        (
            "Achievement unlocked: 'Person Who Has Their Life Together (Financially)'.",
            "— Achievement System",
        ),
        (
            "Your savings goal is complete. Your future self is doing a little dance right now.",
            "— Future You, Thriving",
        ),
        (
            "You saved it all! Now don't spend it all in one place. Or do. We're not your mom.",
            "— SpendWise (Neutral)",
        ),
    ],
    "no_goals": [
        (
            "No savings goals? That's fine. Living on vibes is a valid strategy. (It's not.)",
            "— Financial Advisor, Concerned",
        ),
        (
            "A goal without a plan is just a wish. A wish without savings is just a dream. Add a goal!",
            "— Confucius (Finance Edition)",
        ),
        (
            "You have no savings goals. Your money is just... free range. Wild. Unstructured.",
            "— Chaos Economist",
        ),
        (
            "Set a goal! Even 'save enough to not panic at the grocery store' counts.",
            "— Grocery Store Anxiety Support",
        ),
        (
            "Goals are just dreams with deadlines and spreadsheets. Add one!",
            "— Spreadsheet Enthusiast",
        ),
    ],
}


def _get_motivation_quote(
    monthly_saved: float, goals_data: list, last_quote: str = None
) -> dict:
    """Pick a contextually funny quote based on the user's actual savings state."""
    if not goals_data:
        bucket = "no_goals"
    else:
        best_pct = max(g["progress_pct"] for g in goals_data)
        any_complete = any(g["is_complete"] for g in goals_data)
        if any_complete:
            bucket = "goal_reached"
        elif monthly_saved <= 0:
            bucket = "broke"
        elif best_pct < 15:
            bucket = "just_started"
        elif best_pct < 70:
            bucket = "making_progress"
        else:
            bucket = "almost_there"

    # Filter out the last quote to guarantee variety
    available = [q for q in _MOTIVATION_QUOTES[bucket] if q[0] != last_quote]
    if not available:  # Fallback if somehow all filtered
        available = _MOTIVATION_QUOTES[bucket]

    quote, author = random.choice(available)
    return {"quote": quote, "author": author, "bucket": bucket}


# ── Helpers ───────────────────────────────────────────────
CATEGORY_ICONS = {
    "rent": "🏠",
    "transport": "🚗",
    "health": "❤️",
    "groceries": "🛒",
    "entertainment": "🎬",
    "shopping": "🛍️",
    "food": "🍽️",
    "utilities": "💡",
    "other": "📦",
}

CATEGORY_COLORS = {
    "rent": "#67e8f9",
    "transport": "#c4b5fd",
    "health": "#bbf7d0",
    "groceries": "#fde68a",
    "entertainment": "#fca5a5",
    "shopping": "#fdba74",
    "food": "#a5f3fc",
    "utilities": "#d9f99d",
    "other": "#e2e8f0",
}

CATEGORY_ACCENTS = {
    "rent": {"bg": "#e0f7fa", "fg": "#0097a7"},
    "transport": {"bg": "#ede9ff", "fg": "#7c3aed"},
    "health": {"bg": "#dcfce7", "fg": "#16a34a"},
    "groceries": {"bg": "#fef9c3", "fg": "#a16207"},
    "entertainment": {"bg": "#fee2e2", "fg": "#dc2626"},
    "shopping": {"bg": "#ffedd5", "fg": "#c2410c"},
    "food": {"bg": "#cffafe", "fg": "#0891b2"},
    "utilities": {"bg": "#ecfccb", "fg": "#4d7c0f"},
    "other": {"bg": "#e5e7eb", "fg": "#475569"},
}

INSIGHTS_DONUT_CIRCUMFERENCE = round(2 * pi * 46, 1)
INSIGHTS_SEGMENT_LIMIT = 5
MONTHLY_SCORE_CIRCUMFERENCE = round(2 * pi * 40, 1)
MOTIVATION_BADGES = {
    "start": "Start",
    "reset": "Reset",
    "build": "Build",
    "steady": "Steady",
    "strong": "Strong",
    "elite": "Elite",
    "clean": "Clean",
}


def _motivation_tone(saved_amount: float, expense_amount: float) -> str:
    spent = float(expense_amount or 0)
    saved = float(saved_amount or 0)
    ratio = saved / max(spent, 1)

    if saved > 0 and spent == 0:
        return "clean"
    if saved <= 0 and spent > 0:
        return "reset"
    if saved > 0 and ratio < 0.25:
        return "build"
    if 0.25 <= ratio < 0.75:
        return "steady"
    if 0.75 <= ratio < 1.2:
        return "strong"
    if ratio >= 1.2:
        return "elite"
    return "start"


def _local_ai_style_motivation_quote(
    saved_amount: float, expense_amount: float, tone: str
) -> str:
    intros = {
        "reset": [
            "A reset month is still progress.",
            "You are one decision away from momentum.",
            "This is the perfect point to restart your savings rhythm.",
        ],
        "build": [
            "Nice start, your savings engine is warming up.",
            "You are building traction with every small win.",
            "Momentum is visible, now keep it steady.",
        ],
        "steady": [
            "Your money discipline is getting stronger.",
            "This is balanced money management in action.",
            "You are saving and spending with intention.",
        ],
        "strong": [
            "Powerful month so far, your consistency is working.",
            "You are close to a standout savings month.",
            "This is high-quality financial behavior.",
        ],
        "elite": [
            "Elite control, your savings are leading this month.",
            "You are in wealth-building mode right now.",
            "Outstanding pace, keep protecting this edge.",
        ],
        "clean": [
            "Clean month start, this is premium discipline.",
            "Strong control: savings positive with zero expense.",
            "You started this cycle like a pro saver.",
        ],
        "start": [
            "Every strong saver starts with one clean step.",
            "Today can be the day your saving habit locks in.",
            "Begin small, stay consistent, grow fast.",
        ],
    }
    actions = [
        "Move one non-essential spend into savings today.",
        "Keep one daily spend capped and redirect the rest.",
        "Use a simple 24-hour pause before non-urgent purchases.",
        "Set a tiny daily savings target and defend it.",
        "Track one category tightly for the next 7 days.",
    ]
    payoffs = [
        "Your future self will thank you.",
        "Small discipline now creates big freedom later.",
        "Consistency beats intensity in personal finance.",
        "That habit compounds faster than you think.",
        "This is how stable wealth gets built.",
    ]

    saved = float(saved_amount or 0)
    spent = float(expense_amount or 0)
    seed = int(date.today().strftime("%Y%m%d")) + int(saved * 11) + int(spent * 7)
    rng = random.Random(seed)
    intro_pool = intros.get(tone, intros["start"])

    saved_text = f"₹{abs(saved):,.2f}"
    spent_text = f"₹{abs(spent):,.2f}"
    return (
        f"{rng.choice(intro_pool)} You saved {saved_text} while spending {spent_text}. "
        f"{rng.choice(actions)} {rng.choice(payoffs)}"
    )


def _extract_openai_text(response_payload: dict) -> str:
    output = response_payload.get("output")
    if not isinstance(output, list):
        return ""

    texts = []
    for item in output:
        if not isinstance(item, dict):
            continue
        for content in item.get("content", []):
            if not isinstance(content, dict):
                continue
            text = content.get("text")
            if isinstance(text, str) and text.strip():
                texts.append(text.strip())

    return " ".join(texts).strip()


def _generate_motivation_quote(
    saved_amount: float, expense_amount: float, tone: str
) -> tuple[str, bool]:
    fallback_quote = _local_ai_style_motivation_quote(
        saved_amount, expense_amount, tone
    )
    api_key = getattr(settings, "OPENAI_API_KEY", "").strip()
    if not api_key:
        return fallback_quote, False

    model = (
        getattr(settings, "OPENAI_MOTIVATION_MODEL", "gpt-4.1-mini").strip()
        or "gpt-4.1-mini"
    )
    prompt = (
        "Write one concise motivational saving message for a personal finance app user. "
        f"Saved amount: {saved_amount:.2f} INR. Spent amount: {expense_amount:.2f} INR. "
        f"Tone: {tone}. Keep it under 40 words. Be positive, practical, and premium sounding."
    )
    payload = {
        "model": model,
        "input": prompt,
        "temperature": 0.9,
        "max_output_tokens": 90,
    }
    body = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        "https://api.openai.com/v1/responses",
        data=body,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )

    try:
        with urllib.request.urlopen(req, timeout=10) as response:
            response_payload = json.loads(response.read().decode("utf-8"))
    except (
        urllib.error.URLError,
        urllib.error.HTTPError,
        json.JSONDecodeError,
        TimeoutError,
        ValueError,
    ):
        return fallback_quote, False

    ai_text = _extract_openai_text(response_payload)
    if not ai_text:
        return fallback_quote, False

    clean_text = " ".join(ai_text.split()).strip()
    return clean_text[:260], True


def _serialize_category_tiles(cat_data):
    """Return category data sorted by highest spend for template/API usage."""
    sorted_items = sorted(
        cat_data.items(),
        key=lambda item: item[1]["amount"],
        reverse=True,
    )
    return [
        {
            "key": key,
            "label": value["label"],
            "amount": float(value["amount"]),
            "icon": value["icon"],
            "color": value["color"],
        }
        for key, value in sorted_items
    ]


def _build_insight_segments(cat_tiles, total_expense):
    """Build SVG stroke data for the dashboard insights donut."""
    if total_expense <= 0:
        return []

    dash_offset = 0
    segments = []
    total_expense_float = float(total_expense)

    for tile in cat_tiles[:INSIGHTS_SEGMENT_LIMIT]:
        segment_size = round(
            INSIGHTS_DONUT_CIRCUMFERENCE * (tile["amount"] / total_expense_float),
            1,
        )
        segments.append(
            {
                **tile,
                "dasharray": f"{segment_size} {round(max(INSIGHTS_DONUT_CIRCUMFERENCE - segment_size, 0), 1)}",
                "dashoffset": round(-dash_offset, 1),
                "share_pct": round(tile["amount"] / total_expense_float * 100, 1),
            }
        )
        dash_offset += segment_size

    return segments


def _dashboard_stats_payload(stats):
    """Serialize dashboard stats for JSON responses."""
    return {
        "total_income": float(stats["total_income"]),
        "total_expense": float(stats["total_expense"]),
        "total_saved": float(stats["total_saved"]),
        "total_balance": float(stats["total_balance"]),
        "spend_pct": stats["spend_pct"],
        "savings_pct": stats["savings_pct"],
        "ring_dash": stats["ring_dash"],
        "ring_gap": stats["ring_gap"],
        "cat_tiles": stats["cat_tiles"],
        "insight_segments": stats["insight_segments"],
        "insight_total": float(stats["insight_total"]),
        "insight_stock": stats["insight_stock"],
        "chart_months": stats["chart_months"],
        "coach": get_spending_message(stats["spend_pct"]),
    }


def _chatbot_savings_message(profile, stats, previous_variant=None):
    """Return a short salary-aware nudge after an expense is recorded."""
    salary = Decimal(str(profile.salary or 0))
    spent = Decimal(str(stats["total_expense"]))
    saved = Decimal(str(stats["total_saved"]))

    if salary <= 0:
        choices = [
            "Tiny entries build sharp money habits. Your future wallet noticed.",
            "Expense logged. Your budget brain just earned one experience point.",
            "Tracked money behaves better than mystery money. Nice catch.",
        ]
    else:
        spend_ratio = spent / salary
        remaining_text = indian_currency(max(saved, Decimal("0")))

        if saved < 0:
            over_text = indian_currency(abs(saved))
            choices = [
                f"You are −{over_text} over budget. No judgement — noticing it gives you a way back.",
                f"The balance is −{over_text}. Pause, reset, and let the next small choice help.",
                f"You have crossed the budget by −{over_text}. Tracking it honestly is already progress.",
            ]
        elif spend_ratio >= Decimal("0.9"):
            choices = [
            f"You have {remaining_text} left from this month's salary. Wallet, enter stealth mode.",
            f"Only {remaining_text} remains. Even your cart needs a quiet day.",
            f"{remaining_text} is still standing. Protect it like the last slice of pizza.",
            ]
        elif spend_ratio >= Decimal("0.6"):
            choices = [
            f"{remaining_text} remains. The budget is bending, not breaking.",
            f"You still have {remaining_text}. A no-spend evening would look excellent here.",
            f"{remaining_text} left. Future you votes for keeping some of it.",
            ]
        elif spend_ratio >= Decimal("0.3"):
            choices = [
            f"{remaining_text} remains. Good tracking is doing the heavy lifting.",
            f"You still have {remaining_text}. Your savings plan is very much alive.",
            f"{remaining_text} left. SpendWise approves this level of awareness.",
            ]
        else:
            choices = [
            f"{remaining_text} remains. Strong start, keep the impulse buys guessing.",
            f"You still have {remaining_text}. Your salary can breathe comfortably.",
            f"{remaining_text} left. Future you just sent a tiny thank-you note.",
            ]

    available_variants = [
        index for index in range(len(choices)) if index != previous_variant
    ]
    variant = random.choice(available_variants or list(range(len(choices))))
    return choices[variant], variant


def _serialize_txn(txn: Transaction) -> dict:
    return {
        "id": txn.id,
        "title": txn.title,
        "amount": float(txn.amount),
        "txn_type": txn.txn_type,
        "category": txn.category,
        "category_label": txn.get_category_display(),
        "date": txn.date.isoformat(),
        "icon": CATEGORY_ICONS.get(txn.category, "📦"),
    }


def _available_expense_dates(user) -> list[str]:
    raw_dates = (
        Transaction.objects.filter(user=user, txn_type="expense")
        .order_by("date")
        .values_list("date", flat=True)
        .distinct()
    )
    return [d.isoformat() for d in raw_dates]


def _add_months(month_start: date, delta: int) -> date:
    month_index = month_start.month - 1 + delta
    year = month_start.year + month_index // 12
    month = month_index % 12 + 1
    return date(year, month, 1)


def _month_bounds(month_start: date) -> tuple[date, date]:
    next_month = _add_months(month_start, 1)
    return month_start, next_month - timedelta(days=1)


def _parse_month_param(month_value: str | None) -> date:
    if month_value:
        try:
            year_str, month_str = month_value.split("-", 1)
            return date(int(year_str), int(month_str), 1)
        except (TypeError, ValueError):
            pass
    today = date.today()
    return today.replace(day=1)


def _format_axis_currency(value: float) -> str:
    if value >= 1000:
        display = value / 1000
        suffix = "k"
        digits = 1 if display < 10 and display % 1 else 0
        return f"₹{display:.{digits}f}{suffix}"
    return f"₹{int(round(value))}"


def _pct_change(current_value: Decimal, previous_value: Decimal) -> float:
    current_float = float(current_value)
    previous_float = float(previous_value)

    if previous_float == 0:
        if current_float > 0:
            return 100.0
        if current_float < 0:
            return -100.0
        return 0.0

    return round(((current_float - previous_float) / abs(previous_float)) * 100, 1)


def _build_monthly_weeks(user, month_start: date, month_end: date) -> list[dict]:
    chart_width = 560
    chart_height = 220
    chart_inner_height = 190
    pad_top = chart_height - chart_inner_height  # 30px padding at top
    weeks = []
    cursor = month_start
    week_index = 1

    while cursor <= month_end:
        week_end = min(cursor + timedelta(days=6), month_end)
        income = Transaction.objects.filter(
            user=user,
            txn_type="income",
            date__gte=cursor,
            date__lte=week_end,
        ).aggregate(s=Sum("amount"))["s"] or Decimal("0")
        expense = Transaction.objects.filter(
            user=user,
            txn_type="expense",
            date__gte=cursor,
            date__lte=week_end,
        ).aggregate(s=Sum("amount"))["s"] or Decimal("0")
        weeks.append(
            {
                "label": f"Week {week_index}",
                "range_label": f"{cursor.strftime('%b')} {cursor.day} - {week_end.strftime('%b')} {week_end.day}",
                "income": float(income),
                "expense": float(expense),
            }
        )
        cursor = week_end + timedelta(days=1)
        week_index += 1

    if not weeks:
        weeks.append(
            {
                "label": "Week 1",
                "range_label": f"{month_start.strftime('%b')} {month_start.day} - {month_end.strftime('%b')} {month_end.day}",
                "income": 0.0,
                "expense": 0.0,
            }
        )

    max_val = max(
        max((week["expense"] for week in weeks), default=0),
        1,
    )
    slot_width = chart_width / len(weeks)
    bar_width = min(40, max(24, int(slot_width * 0.55)))

    for index, week in enumerate(weeks):
        # Center the single expense bar in the slot
        center_x = index * slot_width + slot_width / 2
        expense_x = center_x - bar_width / 2

        expense_height = (
            0
            if week["expense"] <= 0
            else max(10, round((week["expense"] / max_val) * chart_inner_height, 1))
        )
        expense_y = pad_top + (chart_inner_height - expense_height)

        week.update(
            {
                "income_height": 0,
                "expense_height": expense_height,
                "income_y": chart_height,
                "expense_y": expense_y,
                "income_x": round(expense_x, 1),
                "expense_x": round(expense_x, 1),
                "bar_width": bar_width,
            }
        )

    trend_path = ""

    axis_step = ceil(max_val / 5) if max_val else 1
    y_axis_labels = [
        _format_axis_currency(axis_step * level) for level in range(5, -1, -1)
    ]

    return {
        "weeks": weeks,
        "trend_path": trend_path,
        "y_axis_labels": y_axis_labels,
        "max_value": round(max_val, 2),
    }


def _build_monthly_categories(
    user, month_start: date, month_end: date, total_expense: Decimal
) -> list[dict]:
    categories = []
    total_expense_float = float(total_expense)

    for key, label in Transaction.CATEGORY_CHOICES:
        amount = Transaction.objects.filter(
            user=user,
            txn_type="expense",
            category=key,
            date__gte=month_start,
            date__lte=month_end,
        ).aggregate(s=Sum("amount"))["s"] or Decimal("0")

        if amount <= 0:
            continue

        accent = CATEGORY_ACCENTS.get(key, CATEGORY_ACCENTS["other"])
        categories.append(
            {
                "key": key,
                "label": label,
                "amount": float(amount),
                "share_pct": round((float(amount) / total_expense_float) * 100, 1)
                if total_expense_float
                else 0,
                "color": CATEGORY_COLORS.get(key, "#e2e8f0"),
                "icon_bg": accent["bg"],
                "icon_fg": accent["fg"],
            }
        )

    return sorted(categories, key=lambda category: category["amount"], reverse=True)


def _build_monthly_score(
    total_income: Decimal, total_expense: Decimal, total_saved: Decimal, salary: Decimal
) -> dict:
    if salary > 0:
        baseline = float(salary)
    else:
        baseline = float(total_income)

    if baseline > 0:
        spend_ratio = float(total_expense) / baseline
        save_ratio = max(float(total_saved), 0) / baseline
        score = round(
            max(0, min(100, 100 - (spend_ratio * 62) + min(save_ratio, 1) * 24))
        )
    else:
        score = 0

    if score >= 80:
        title = "Excellent rhythm"
        description = (
            "You're keeping spending under control and protecting your savings."
        )
    elif score >= 60:
        title = "Good standing"
        description = "Your month looks healthy, with room to tighten a few categories."
    elif score >= 40:
        title = "Needs attention"
        description = "Spending is starting to crowd out savings this month."
    else:
        title = "High pressure"
        description = (
            "This month is running hot. Review your biggest expense days first."
        )

    return {
        "value": score,
        "ring_dash": round((score / 100) * MONTHLY_SCORE_CIRCUMFERENCE, 1),
        "ring_gap": round(
            MONTHLY_SCORE_CIRCUMFERENCE - (score / 100) * MONTHLY_SCORE_CIRCUMFERENCE, 1
        ),
        "title": title,
        "description": description,
    }


def _build_top_spending_days(user, month_start: date, month_end: date) -> list[dict]:
    daily_spend = {}
    badge_classes = ["payment", "figma", "withdrawal", "webflow", "zalando"]

    for txn in Transaction.objects.filter(
        user=user,
        txn_type="expense",
        date__gte=month_start,
        date__lte=month_end,
    ).order_by("date", "created_at"):
        entry = daily_spend.setdefault(
            txn.date,
            {
                "amount": Decimal("0"),
                "categories": {},
            },
        )
        entry["amount"] += txn.amount
        entry["categories"][txn.category] = (
            entry["categories"].get(txn.category, Decimal("0")) + txn.amount
        )

    ranked_days = sorted(
        daily_spend.items(), key=lambda item: item[1]["amount"], reverse=True
    )[:5]
    results = []

    for index, (day, payload) in enumerate(ranked_days):
        top_categories = sorted(
            payload["categories"].items(), key=lambda item: item[1], reverse=True
        )[:2]
        category_text = " + ".join(
            dict(Transaction.CATEGORY_CHOICES).get(key, key.title())
            for key, _ in top_categories
        )
        results.append(
            {
                "day_number": day.day,
                "label": day.strftime("%b %d"),
                "summary": category_text or "Expenses",
                "amount": float(payload["amount"]),
                "badge_class": badge_classes[index % len(badge_classes)],
            }
        )

    return results


def _build_monthly_analysis(user, profile, selected_month: date) -> dict:
    month_start, month_end = _month_bounds(selected_month)
    prev_month_start = _add_months(month_start, -1)
    prev_month_end = month_start - timedelta(days=1)
    current_month_str = month_start.strftime("%Y-%m")

    month_txns = Transaction.objects.filter(
        user=user, date__gte=month_start, date__lte=month_end
    )
    total_income = month_txns.filter(txn_type="income").aggregate(s=Sum("amount"))[
        "s"
    ] or Decimal("0")
    total_expense = month_txns.filter(txn_type="expense").aggregate(s=Sum("amount"))[
        "s"
    ] or Decimal("0")
    
    # Check for excess income for this month
    from .models import ExcessIncome
    excess_income_obj = ExcessIncome.objects.filter(user=user, month=current_month_str).first()
    excess_income = Decimal(str(excess_income_obj.amount)) if excess_income_obj else Decimal("0")
    
    salary = Decimal(str(profile.salary)) if profile.salary else Decimal("0")
    # Include excess income in savings calculation
    if salary > 0:
        total_income_with_excess = salary + excess_income
        total_saved = total_income_with_excess - total_expense
    else:
        total_saved = total_income - total_expense
    total_balance = total_saved

    prev_txns = Transaction.objects.filter(
        user=user, date__gte=prev_month_start, date__lte=prev_month_end
    )
    prev_income = prev_txns.filter(txn_type="income").aggregate(s=Sum("amount"))[
        "s"
    ] or Decimal("0")
    prev_expense = prev_txns.filter(txn_type="expense").aggregate(s=Sum("amount"))[
        "s"
    ] or Decimal("0")
    
    # Check for excess income for previous month
    prev_month_str = prev_month_start.strftime("%Y-%m")
    prev_excess_income_obj = ExcessIncome.objects.filter(user=user, month=prev_month_str).first()
    prev_excess_income = Decimal(str(prev_excess_income_obj.amount)) if prev_excess_income_obj else Decimal("0")
    
    if salary > 0:
        prev_total_income_with_excess = salary + prev_excess_income
        prev_saved = prev_total_income_with_excess - prev_expense
    else:
        prev_saved = prev_income - prev_expense

    balance_change_pct = _pct_change(total_balance, prev_saved)
    categories = _build_monthly_categories(user, month_start, month_end, total_expense)
    donut_segments = _build_insight_segments(categories[:4], total_expense)
    weekly_chart = _build_monthly_weeks(user, month_start, month_end)
    monthly_score = _build_monthly_score(
        total_income, total_expense, total_saved, salary
    )
    top_spending_days = _build_top_spending_days(user, month_start, month_end)

    # Calculate spend ratio including excess income
    if salary > 0:
        total_income_for_calc = float(salary + excess_income)
        spend_ratio = min(round((float(total_expense) / total_income_for_calc) * 100, 1), 999.9) if total_income_for_calc > 0 else 0.0
    elif total_income > 0:
        spend_ratio = min(
            round((float(total_expense) / float(total_income)) * 100, 1), 999.9
        )
    else:
        spend_ratio = 0.0

    # Get highest priority goal (if multiple high priority, pick the one with lowest progress)
    from .models import SavingsGoal
    
    high_priority_goals = SavingsGoal.objects.filter(
        user=user, 
        priority='high',
        is_active=True
    ).order_by('saved_amount')  # Lowest saved amount first (needs most attention)
    
    if high_priority_goals.exists():
        priority_goal = high_priority_goals.first()
        goal_progress_pct = (
            min(round((float(priority_goal.saved_amount) / float(priority_goal.target_amount)) * 100, 1), 100)
            if priority_goal.target_amount > 0
            else 0
        )
        wealth_title = priority_goal.name
        wealth_amount_label = f"₹{priority_goal.target_amount:,.2f} target"
        wealth_progress_pct = goal_progress_pct
        wealth_sub = f"{goal_progress_pct:.1f}% of your goal reached" if priority_goal.target_amount > 0 else "Set a target"
        wealth_badge = "HIGH PRIORITY"
        wealth_cta = "KEEP PUSHING TOWARD YOUR GOAL" if goal_progress_pct < 100 else "GOAL ACHIEVED!"
    else:
        # Fallback to old wealth/target savings behavior
        target = (
            Decimal(str(profile.target_savings)) if profile.target_savings else Decimal("0")
        )
        wealth_progress_pct = (
            min(round((max(float(total_saved), 0) / float(target)) * 100, 1), 100)
            if target > 0
            else 0
        )
        wealth_title = "Wealth"
        wealth_amount_label = f"₹{target:,.2f} target" if target > 0 else "No goal set"
        wealth_sub = (
            f"{wealth_progress_pct:.1f}% of your savings goal reached"
            if target > 0
            else "Set a savings target to track progress here"
        )
        wealth_badge = "GOALS" if target > 0 else "START"
        wealth_cta = (
            "BUILD YOUR GOAL CONSISTENCY"
            if target > 0
            else "SET A TARGET TO TRACK PROGRESS"
        )

    if total_expense <= 0:
        reminder_title = "Start tracking expenses for this month"
        reminder_sub = "ADD A FEW TRANSACTIONS TO SEE TRENDS"
    elif monthly_score["value"] >= 75:
        reminder_title = "Your monthly budget is holding up well"
        reminder_sub = "KEEP THIS MOMENTUM GOING"
    else:
        reminder_title = "Review your biggest spending days this month"
        reminder_sub = "CUT BACK WHERE IT HURTS MOST"

    return {
        "selected_month": month_start,
        "selected_month_param": month_start.strftime("%Y-%m"),
        "selected_month_label": month_start.strftime("%B %Y"),
        "month_period_label": f"{month_start.day} - {month_end.day} {month_end.strftime('%b %Y')}",
        "prev_month_param": _add_months(month_start, -1).strftime("%Y-%m"),
        "next_month_param": _add_months(month_start, 1).strftime("%Y-%m"),
        "is_current_month": month_start == date.today().replace(day=1),
        "account_subtitle": f"{month_txns.count()} transaction{'s' if month_txns.count() != 1 else ''} in this month",
        "month_badge": "This month"
        if month_start == date.today().replace(day=1)
        else month_start.strftime("%b %Y"),
        "total_income": total_income,
        "total_expense": total_expense,
        "total_saved": total_saved,
        "total_balance": total_balance,
        "excess_income": excess_income,
        "has_excess_income": excess_income > 0,
        "balance_change_pct": balance_change_pct,
        "balance_change_positive": balance_change_pct >= 0,
        "weekly_chart": weekly_chart,
        "categories": categories[:5],
        "donut_segments": donut_segments,
        "donut_legend": categories[:4],
        "monthly_score": monthly_score,
        "top_spending_days": top_spending_days,
        "reminder_title": reminder_title,
        "reminder_sub": reminder_sub,
        "wealth_title": wealth_title,
        "wealth_amount_label": wealth_amount_label,
        "wealth_progress_pct": wealth_progress_pct,
        "wealth_sub": wealth_sub,
        "wealth_badge": wealth_badge,
        "wealth_cta": wealth_cta,
        "spend_ratio_pct": spend_ratio,
    }


def _build_monthly_analysis_email_context(user, profile, analysis: dict) -> dict:
    weekly_rows = [
        {
            "label": week["range_label"],
            "expense": indian_currency(week["expense"]),
        }
        for week in analysis["weekly_chart"]["weeks"]
        if float(week["expense"]) > 0
    ]
    top_days = [
        {
            "label": day["label"],
            "summary": day["summary"],
            "amount": indian_currency(day["amount"]),
        }
        for day in analysis["top_spending_days"]
    ]
    categories = [
        {
            "label": category["label"],
            "amount": indian_currency(category["amount"]),
            "share_pct": category["share_pct"],
        }
        for category in analysis["categories"]
    ]

    return {
        "user": user,
        "profile": profile,
        "selected_month_label": analysis["selected_month_label"],
        "month_period_label": analysis["month_period_label"],
        "account_subtitle": analysis["account_subtitle"],
        "total_income": indian_currency(analysis["total_income"]),
        "salary": indian_currency(profile.salary) if profile.salary else "Not set",
        "excess_income": indian_currency(analysis["excess_income"]),
        "has_excess_income": analysis["has_excess_income"],
        "total_expense": indian_currency(analysis["total_expense"]),
        "total_saved": indian_currency(analysis["total_saved"], True),
        "total_balance": indian_currency(analysis["total_balance"]),
        "balance_change_pct": analysis["balance_change_pct"],
        "balance_change_positive": analysis["balance_change_positive"],
        "spend_ratio_pct": analysis["spend_ratio_pct"],
        "monthly_score": analysis["monthly_score"],
        "reminder_title": analysis["reminder_title"],
        "reminder_sub": analysis["reminder_sub"],
        "wealth_title": analysis["wealth_title"],
        "wealth_amount_label": analysis["wealth_amount_label"],
        "wealth_badge": analysis["wealth_badge"],
        "wealth_cta": analysis["wealth_cta"],
        "wealth_progress_pct": analysis["wealth_progress_pct"],
        "wealth_sub": analysis["wealth_sub"],
        "categories": categories,
        "weekly_rows": weekly_rows,
        "top_days": top_days,
    }





def _build_insight_stock_card(chart_months, total_expense):
    """Create stock-style summary data for the dashboard hero card."""
    latest_expense = float(total_expense)
    previous_expense = chart_months[-2]["expense"] if len(chart_months) > 1 else 0

    if previous_expense > 0:
        change_pct = round(
            ((latest_expense - previous_expense) / previous_expense) * 100, 1
        )
    elif latest_expense > 0:
        change_pct = 100.0
    else:
        change_pct = 0.0

    return {
        "title": "Expense Trend",
        "symbol": "SPND",
        "subtitle": "Monthly spend",
        "amount": latest_expense,
        "change_pct": abs(change_pct),
        "is_up": change_pct < 0,  # expenses DOWN = good = green arrow up
    }


def _dashboard_stats(user, profile):
    """Compute all dynamic stats for the dashboard."""
    today = date.today()
    month_start = today.replace(day=1)
    current_month_str = month_start.strftime("%Y-%m")

    txns = Transaction.objects.filter(user=user, date__gte=month_start, date__lte=today)

    total_income = txns.filter(txn_type="income").aggregate(s=Sum("amount"))[
        "s"
    ] or Decimal("0")
    total_expense = txns.filter(txn_type="expense").aggregate(s=Sum("amount"))[
        "s"
    ] or Decimal("0")

    # Check for excess income this month
    from .models import ExcessIncome
    excess_income_obj = ExcessIncome.objects.filter(user=user, month=current_month_str).first()
    excess_income = Decimal(str(excess_income_obj.amount)) if excess_income_obj else Decimal("0")

    # Savings = (salary + excess_income) - expenses (if salary set), else income - expenses
    salary = Decimal(str(profile.salary)) if profile.salary else Decimal("0")
    if salary > 0:
        total_income_with_excess = salary + excess_income
        total_saved = total_income_with_excess - total_expense
        total_balance = total_income_with_excess - total_expense
    else:
        total_saved = total_income - total_expense
        total_balance = total_income - total_expense

    # Category breakdown (expenses only)
    cat_data = {}
    for cat, label in Transaction.CATEGORY_CHOICES:
        amt = txns.filter(txn_type="expense", category=cat).aggregate(s=Sum("amount"))[
            "s"
        ] or Decimal("0")
        if amt > 0:
            cat_data[cat] = {
                "label": label,
                "amount": amt,
                "icon": CATEGORY_ICONS.get(cat, "📦"),
                "color": CATEGORY_COLORS.get(cat, "#e2e8f0"),
            }
    cat_tiles = _serialize_category_tiles(cat_data)
    insight_segments = _build_insight_segments(cat_tiles, total_expense)

    # Spend % vs salary (including excess income)
    spend_pct = None
    if salary > 0:
        total_income_for_calc = float(salary + excess_income)
        spend_pct = min(
            round(float(total_expense) / total_income_for_calc * 100, 1), 100
        ) if total_income_for_calc > 0 else 0

    # Savings % vs target
    savings_pct = None
    CIRC = 263.9
    if profile.target_savings and float(profile.target_savings) > 0:
        savings_pct = min(
            round(float(total_saved) / float(profile.target_savings) * 100, 1), 100
        )
    ring_dash = round((savings_pct or 0) / 100 * CIRC, 1)

    # Last 7 transactions
    recent_txns = Transaction.objects.filter(user=user).order_by(
        "-date", "-created_at"
    )[:7]

    # Chart: last 6 months saved vs expenses
    chart_months = []
    for i in range(5, -1, -1):
        d = today.replace(day=1) - timedelta(days=i * 28)
        m_start = d.replace(day=1)
        if d.month == 12:
            m_end = d.replace(year=d.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            m_end = d.replace(month=d.month + 1, day=1) - timedelta(days=1)

        m_exp = Transaction.objects.filter(
            user=user, txn_type="expense", date__gte=m_start, date__lte=m_end
        ).aggregate(s=Sum("amount"))["s"] or Decimal("0")
        m_inc = Transaction.objects.filter(
            user=user, txn_type="income", date__gte=m_start, date__lte=m_end
        ).aggregate(s=Sum("amount"))["s"] or Decimal("0")
        m_saved = max(m_inc - m_exp, Decimal("0"))

        chart_months.append(
            {
                "label": d.strftime("%b"),
                "expense": float(m_exp),
                "saved": float(m_saved),
            }
        )

    # Daily spending for current month
    daily_spending = []
    current_day = month_start
    while current_day <= today:
        day_expense = Transaction.objects.filter(
            user=user,
            txn_type="expense",
            date=current_day
        ).aggregate(s=Sum("amount"))["s"] or Decimal("0")
        
        daily_spending.append({
            "day": current_day.day,
            "date": current_day.isoformat(),
            "amount": float(day_expense),
        })
        current_day += timedelta(days=1)

    insight_stock = _build_insight_stock_card(chart_months, total_expense)

    return {
        "total_income": total_income,
        "total_expense": total_expense,
        "total_saved": total_saved,
        "total_balance": total_balance,
        "excess_income": excess_income,
        "has_excess_income": excess_income > 0,
        "cat_data": cat_data,
        "cat_tiles": cat_tiles,
        "spend_pct": spend_pct,
        "savings_pct": savings_pct,
        "ring_dash": ring_dash,
        "ring_gap": round(CIRC - ring_dash, 1),
        "insight_segments": insight_segments,
        "insight_total": total_expense,
        "insight_stock": insight_stock,
        "recent_txns": recent_txns,
        "chart_months": chart_months,
        "daily_spending": daily_spending,
    }


# ── Onboarding ────────────────────────────────────────────
def onboarding(request: HttpRequest) -> HttpResponse:
    if request.user.is_authenticated:
        return redirect("dashboard")
    return render(request, "login/onboarding.html")


def _issue_signup_otp(profile: UserProfile) -> str:
    code = f"{random.randint(100000, 999999)}"
    profile.email_verification_code = code
    profile.email_verification_sent_at = timezone.now()
    profile.email_is_verified = False
    profile.save(
        update_fields=[
            "email_verification_code",
            "email_verification_sent_at",
            "email_is_verified",
        ]
    )
    return code


def _send_signup_otp_email(user: User, code: str) -> None:
    expiry_minutes = getattr(settings, "EMAIL_VERIFICATION_CODE_EXPIRY_MINUTES", 10)
    name = user.first_name or user.username or "there"
    body = (
        f"Hi {name},\n\n"
        f"Use this 6-digit OTP to verify your SpendWise account: {code}\n\n"
        f"This OTP expires in {expiry_minutes} minutes.\n\n"
        "If you did not create this account, you can ignore this email."
    )
    email = EmailMultiAlternatives(
        subject="Verify your SpendWise account",
        body=body,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[user.email],
    )
    email.send(fail_silently=False)


# ── Sign Up ───────────────────────────────────────────────
def signup(request: HttpRequest) -> HttpResponse:
    if request.user.is_authenticated:
        return redirect("dashboard")

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        email = request.POST.get("email", "").strip().lower()
        password = request.POST.get("password", "")
        confirm_password = request.POST.get("confirm_password", "")
        existing_user = User.objects.filter(email=email).first() if email else None

        errors = {}
        if not name:
            errors["name"] = "Name is required."
        if not email:
            errors["email"] = "Email is required."
        else:
            try:
                validate_email(email)
            except ValidationError:
                errors["email"] = "Enter a valid email address."
            if email and not errors.get("email"):
                if existing_user and existing_user.is_active:
                    errors["email"] = "An account with this email already exists."
        if not password:
            errors["password"] = "Password is required."
        else:
            try:
                validate_password(password)
            except ValidationError as exc:
                errors["password"] = exc.messages[0]
        if not confirm_password:
            errors["confirm_password"] = "Please confirm your password."
        elif password and password != confirm_password:
            errors["confirm_password"] = "Passwords do not match."

        if errors:
            return render(
                request, "login/signup.html", {"errors": errors, "form": request.POST}
            )

        if existing_user and not existing_user.is_active:
            user = existing_user
            user.username = email
            user.email = email
            user.first_name = name.split()[0]
            user.last_name = " ".join(name.split()[1:]) if len(name.split()) > 1 else ""
            user.set_password(password)
            user.is_active = True
            user.save(
                update_fields=[
                    "username",
                    "email",
                    "first_name",
                    "last_name",
                    "password",
                    "is_active",
                ]
            )
            profile, _ = UserProfile.objects.get_or_create(user=user)
        else:
            username = email
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password,
                first_name=name.split()[0],
                last_name=" ".join(name.split()[1:]) if len(name.split()) > 1 else "",
                is_active=True,
            )
            profile = UserProfile.objects.create(user=user)

        # OTP signup is paused for now. To re-enable it, make users inactive,
        # call _issue_signup_otp(profile), send _send_signup_otp_email(...),
        # store pending_signup_user_id in session, then redirect to signup_verify.
        profile.email_is_verified = True
        profile.email_verification_code = ""
        profile.save(update_fields=["email_is_verified", "email_verification_code"])
        login(request, user)
        messages.success(
            request,
            f"Welcome to SpendWise, {user.first_name}! Your account is ready.",
        )
        return redirect("dashboard")

    return render(request, "login/signup.html")


def signup_verify(request: HttpRequest) -> HttpResponse:
    if request.user.is_authenticated:
        return redirect("dashboard")

    pending_user_id = request.session.get("pending_signup_user_id")
    user = User.objects.filter(id=pending_user_id, is_active=False).first()
    if not user:
        messages.error(request, "Your verification session expired. Please sign up again.")
        return redirect("signup")

    profile, _ = UserProfile.objects.get_or_create(user=user)
    errors = {}

    if request.method == "POST":
        action = request.POST.get("action", "verify")

        if action == "resend":
            code = _issue_signup_otp(profile)
            try:
                _send_signup_otp_email(user, code)
                messages.success(request, f"A fresh OTP was sent to {user.email}.")
            except Exception:
                errors["general"] = "We could not resend the OTP. Please check SMTP settings."
            return render(
                request,
                "login/signup_verify.html",
                {"errors": errors, "email": user.email},
            )

        code = request.POST.get("otp", "").strip()
        expiry_minutes = getattr(settings, "EMAIL_VERIFICATION_CODE_EXPIRY_MINUTES", 10)
        sent_at = profile.email_verification_sent_at

        if not code:
            errors["otp"] = "Enter the 6-digit OTP."
        elif code != profile.email_verification_code:
            errors["otp"] = "That OTP is incorrect."
        elif not sent_at or timezone.now() > sent_at + timedelta(minutes=expiry_minutes):
            errors["otp"] = "That OTP expired. Please resend a new one."

        if not errors:
            profile.email_is_verified = True
            profile.email_verification_code = ""
            profile.save(
                update_fields=[
                    "email_is_verified",
                    "email_verification_code",
                ]
            )
            user.is_active = True
            user.save(update_fields=["is_active"])
            request.session.pop("pending_signup_user_id", None)
            login(request, user)
            messages.success(
                request,
                f"Welcome to SpendWise, {user.first_name}! Your account is verified.",
            )
            return redirect("dashboard")

    return render(request, "login/signup.html")




# ── Login ─────────────────────────────────────────────────
def login_view(request: HttpRequest) -> HttpResponse:
    if request.user.is_authenticated:
        return redirect("dashboard")

    if request.method == "POST":
        email = request.POST.get("email", "").strip().lower()
        password = request.POST.get("password", "")

        errors = {}
        if not email:
            errors["email"] = "Email is required."
        else:
            try:
                validate_email(email)
            except ValidationError:
                errors["email"] = "Enter a valid email address."
        if not password:
            errors["password"] = "Password is required."

        if not errors:
            user = authenticate(request, username=email, password=password)
            if user is not None:
                login(request, user)
                return redirect("dashboard")
            else:
                errors["general"] = "Invalid email or password."

        return render(
            request, "login/login.html", {"errors": errors, "form": request.POST}
        )

    return render(request, "login/login.html")


# ── Logout ────────────────────────────────────────────────
def logout_view(request: HttpRequest) -> HttpResponse:
    if request.method != "POST":
        return redirect("dashboard" if request.user.is_authenticated else "onboarding")
    logout(request)
    return redirect("onboarding")


# ── Dashboard ─────────────────────────────────────────────
@login_required(login_url="/login/")
def dashboard(request: HttpRequest) -> HttpResponse:
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    stats = _dashboard_stats(request.user, profile)
    coach = get_spending_message(stats["spend_pct"])
    dad_joke = fetch_dad_joke()

    # Smart motivation quote — context-aware, no repeats
    goals = SavingsGoal.objects.filter(user=request.user)
    monthly_saved = float(stats["total_saved"])
    goals_data = []
    for g in goals:
        if float(g.target_amount) > 0:
            pct = min(
                round(float(monthly_saved) / float(g.target_amount) * 100, 1), 100
            )
        else:
            pct = 0
        pct = max(pct, 0)
        goals_data.append(
            {
                "progress_pct": pct,
                "is_complete": float(monthly_saved) >= float(g.target_amount),
            }
        )

    # Get motivation quote - use API quotes for real motivational content
    api_quote = fetch_motivation()
    motivation = {
        "quote": api_quote.get("text", ""),
        "author": api_quote.get("author", "Unknown"),
        "bucket": "making_progress",  # default bucket
    }

    # If API fails, fallback to context-aware quote
    if not motivation["quote"]:
        last_quote = request.session.get("last_dash_motivation_quote", None)
        motivation = _get_motivation_quote(monthly_saved, goals_data, last_quote)
        request.session["last_dash_motivation_quote"] = motivation["quote"]

    return render(
        request,
        "login/dashboard.html",
        {
            "user": request.user,
            "profile": profile,
            "coach": coach,
            "dad_joke": dad_joke,
            "motivation": motivation,
            "categories": Transaction.CATEGORY_CHOICES,
            "today": date.today().isoformat(),
            "active_nav": "dashboard",
            **stats,
            "chart_months": json.dumps(stats["chart_months"]),
            "insight_segments_json": json.dumps(stats["insight_segments"]),
        },
    )


@login_required(login_url="/login/")
@require_http_methods(["GET"])
def api_dashboard_summary(request: HttpRequest) -> JsonResponse:
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    stats = _dashboard_stats(request.user, profile)
    recent_expenses = [
        _serialize_txn(txn) for txn in stats["recent_txns"] if txn.txn_type == "expense"
    ]
    return JsonResponse(
        {
            "success": True,
            "dad_joke": fetch_dad_joke(),
            "available_expense_dates": _available_expense_dates(request.user),
            "recent_expenses": recent_expenses,
            **_dashboard_stats_payload(stats),
        }
    )


@login_required(login_url="/login/")
@require_http_methods(["GET"])
def api_expenses_by_date(request: HttpRequest) -> JsonResponse:
    requested_date = (request.GET.get("date") or "").strip()
    selected_date = None
    if requested_date:
        try:
            selected_date = date.fromisoformat(requested_date)
        except ValueError:
            return JsonResponse(
                {"error": "Invalid date format. Use YYYY-MM-DD."}, status=400
            )

    qs = Transaction.objects.filter(user=request.user, txn_type="expense").order_by(
        "-date", "-created_at"
    )
    if selected_date:
        qs = qs.filter(date=selected_date)

    expenses = [_serialize_txn(txn) for txn in qs[:30]]
    return JsonResponse(
        {
            "success": True,
            "selected_date": selected_date.isoformat() if selected_date else "",
            "available_expense_dates": _available_expense_dates(request.user),
            "expenses": expenses,
        }
    )


@login_required(login_url="/login/")
@require_http_methods(["GET"])
def api_dad_joke(request: HttpRequest) -> JsonResponse:
    return JsonResponse(
        {
            "success": True,
            "dad_joke": fetch_dad_joke(),
        }
    )


# ── API: Get Motivational Quote ───────────────────────────
@login_required(login_url="/login/")
@require_http_methods(["GET"])
def api_motivation_quote(request: HttpRequest) -> JsonResponse:
    quote = fetch_motivation()
    return JsonResponse(
        {
            "success": True,
            "text": quote.get("text", ""),
            "author": quote.get("author", "Unknown"),
        }
    )


# ── Savings Goals Page ────────────────────────────────────
@login_required(login_url="/login/")
def savings(request: HttpRequest) -> HttpResponse:
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    stats = _dashboard_stats(request.user, profile)
    monthly_saved = stats["total_saved"]  # income - expenses this month
    goals = SavingsGoal.objects.filter(user=request.user)

    # Calculate this month's expenses for allocation
    from datetime import date
    from django.db.models import Sum

    first_of_month = date.today().replace(day=1)
    monthly_expenses = (
        Transaction.objects.filter(
            user=request.user, txn_type="expense", date__gte=first_of_month
        ).aggregate(total=Sum("amount"))["total"]
        or 0
    )
    monthly_expenses = float(monthly_expenses)

    # Calculate available for goals (including excess income)
    from .models import ExcessIncome
    current_month_str = first_of_month.strftime("%Y-%m")
    excess_income_obj = ExcessIncome.objects.filter(user=request.user, month=current_month_str).first()
    excess_income = float(excess_income_obj.amount) if excess_income_obj else 0
    
    salary = float(profile.salary) if profile.salary else 0
    if salary > 0:
        net_available = (salary + excess_income) - monthly_expenses
    else:
        # Fall back to actual income transactions this month
        monthly_income = (
            Transaction.objects.filter(
                user=request.user, txn_type="income", date__gte=first_of_month
            ).aggregate(total=Sum("amount"))["total"]
            or 0
        )
        net_available = float(monthly_income) - monthly_expenses
    available_for_goals = max(net_available, 0)

    # Calculate allocation breakdown
    goals_list = list(goals)
    for g in goals_list:
        if g.last_allocated_month != current_month_str:
            base_saved = float(g.saved_amount)
            g.saved_amount = Decimal("0")
            g.current_month_auto_allocation = Decimal(str(round(base_saved, 2)))
            g.save(update_fields=["saved_amount", "current_month_auto_allocation"])

    available_after_prior = available_for_goals
    for g in goals_list:
        available_after_prior -= float(g.current_month_auto_allocation)
    available_after_prior = max(available_after_prior, 0)

    allocation_dict = _calculate_goal_allocations(available_after_prior, goals_list)

    for g in goals_list:
        allocated = allocation_dict.get(g.id, 0)
        if allocated > 0:
            effective_saved = float(g.saved_amount) + float(g.current_month_auto_allocation)
            new_effective = min(effective_saved + allocated, float(g.target_amount))
            new_auto = new_effective - float(g.saved_amount)
            g.current_month_auto_allocation = Decimal(str(round(new_auto, 2)))
            g.last_allocated_month = current_month_str
            g.save(update_fields=["current_month_auto_allocation", "last_allocated_month"])

    allocation_breakdown = []
    for g in goals_list:
        effective_saved = float(g.saved_amount) + float(g.current_month_auto_allocation)
        if g.is_active:
            allocation_breakdown.append(
                {
                    "name": g.name,
                    "priority": g.priority,
                    "allocation_percentage": g.allocation_percentage,
                    "allocated": allocation_dict.get(g.id, 0),
                }
            )

    # For each goal, progress = effective_saved / target_amount * 100
    CIRC = 263.9
    goals_data = []
    
    for g in goals_list:
        effective_saved = float(g.saved_amount) + float(g.current_month_auto_allocation)
        max_from_available = available_for_goals
        effective_saved = min(effective_saved, float(g.saved_amount) + max_from_available)
        remaining = max(float(g.target_amount) - effective_saved, 0)
        is_complete = effective_saved >= float(g.target_amount)
        if float(g.target_amount) > 0:
            pct = min(round(effective_saved / float(g.target_amount) * 100, 1), 100)
        else:
            pct = 0
        pct = max(pct, 0)
        goals_data.append(
            {
                "id": g.id,
                "name": g.name,
                "target_amount": g.target_amount,
                "saved_amount": effective_saved,
                "progress_pct": pct,
                "ring_dash": round(pct / 100 * CIRC, 1),
                "ring_gap": round(CIRC - pct / 100 * CIRC, 1),
                "is_complete": is_complete,
                "is_active": g.is_active and not is_complete,
                "remaining": remaining,
                "priority": g.priority,
                "allocation_percentage": g.allocation_percentage,
            }
        )

    # Get motivation quote - use API quotes for real motivational content
    api_quote = fetch_motivation()
    motivation = {
        "quote": api_quote.get("text", ""),
        "author": api_quote.get("author", "Unknown"),
        "bucket": "making_progress",  # default bucket
    }

    # If API fails, fallback to context-aware quote
    if not motivation["quote"]:
        last_quote = request.session.get("last_motivation_quote", None)
        motivation = _get_motivation_quote(float(monthly_saved), goals_data, last_quote)
        request.session["last_motivation_quote"] = motivation["quote"]

    return render(
        request,
        "login/savings.html",
        {
            "user": request.user,
            "profile": profile,
            "goals": goals_data,
            "goal_count": goals.count(),
            "monthly_saved": monthly_saved,
            "monthly_expenses": monthly_expenses,
            "available_for_goals": net_available,
            "allocation_breakdown": allocation_breakdown,
            "motivation": motivation,
            "active_nav": "savings",
        },
    )


# ── API: Create Goal ──────────────────────────────────────
@login_required(login_url="/login/")
@require_POST
def api_create_goal(request: HttpRequest) -> JsonResponse:
    try:
        data = json.loads(request.body)
        name = data.get("name", "").strip() or "My Goal"
        target = float(data.get("target_amount", 0) or 0)
        priority = data.get("priority", "medium")
        allocation_pct = int(data.get("allocation_percentage", 50))

        # Clamp percentage to valid range
        allocation_pct = max(0, min(100, allocation_pct))

        # Validate priority
        valid_priorities = ["high", "medium", "low"]
        if priority not in valid_priorities:
            priority = "medium"

        if target <= 0:
            return JsonResponse(
                {"error": "Target must be greater than zero."}, status=400
            )
        goal = SavingsGoal.objects.create(
            user=request.user,
            name=name,
            target_amount=Decimal(str(target)),
            saved_amount=Decimal("0"),
            priority=priority,
            allocation_percentage=allocation_pct,
            is_active=True,
        )
        return JsonResponse({"success": True, "goal": _goal_json(goal)})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# ── API: Update Goal ──────────────────────────────────────
@login_required(login_url="/login/")
@require_http_methods(["PUT"])
def api_update_goal(request: HttpRequest, goal_id: int) -> JsonResponse:
    try:
        goal = SavingsGoal.objects.get(id=goal_id, user=request.user)
        data = json.loads(request.body)
        goal.name = data.get("name", goal.name).strip() or goal.name
        goal.target_amount = Decimal(
            str(float(data.get("target_amount", goal.target_amount)))
        )

        # Update priority if provided
        priority = data.get("priority")
        if priority and priority in ["high", "medium", "low"]:
            goal.priority = priority

        # Update allocation percentage if provided
        allocation_pct = data.get("allocation_percentage")
        if allocation_pct is not None:
            goal.allocation_percentage = max(0, min(100, int(allocation_pct)))

        # Update is_active if provided
        is_active = data.get("is_active")
        if is_active is not None:
            goal.is_active = bool(is_active)

        goal.save()
        return JsonResponse({"success": True, "goal": _goal_json(goal)})
    except SavingsGoal.DoesNotExist:
        return JsonResponse({"error": "Not found."}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# ── API: Delete Goal ──────────────────────────────────────
@login_required(login_url="/login/")
@require_http_methods(["DELETE"])
def api_delete_goal(request: HttpRequest, goal_id: int) -> JsonResponse:
    try:
        goal = SavingsGoal.objects.get(id=goal_id, user=request.user)
        goal.delete()
        return JsonResponse({"success": True})
    except SavingsGoal.DoesNotExist:
        return JsonResponse({"error": "Not found."}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# ── API: Contribute to Goal ───────────────────────────────
@login_required(login_url="/login/")
@require_POST
def api_contribute_goal(request: HttpRequest, goal_id: int) -> JsonResponse:
    try:
        goal = SavingsGoal.objects.get(id=goal_id, user=request.user)
        data = json.loads(request.body)
        amount = float(data.get("amount", 0) or 0)
        if amount <= 0:
            return JsonResponse(
                {"error": "Amount must be greater than zero."}, status=400
            )
        goal.saved_amount += Decimal(str(amount))

        # Auto-mark as complete if target reached
        if goal.saved_amount >= goal.target_amount:
            goal.is_active = False

        goal.save()
        return JsonResponse({"success": True, "goal": _goal_json(goal)})
    except SavingsGoal.DoesNotExist:
        return JsonResponse({"error": "Not found."}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# ── API: Get Goal Allocation Breakdown ─────────────────────
@login_required(login_url="/login/")
@require_GET
def api_goal_allocations(request: HttpRequest) -> JsonResponse:
    try:
        profile, _ = UserProfile.objects.get_or_create(user=request.user)
        salary = float(profile.salary) if profile.salary else 0

        from datetime import date
        from django.db.models import Sum
        from decimal import Decimal
        from .models import ExcessIncome

        first_of_month = date.today().replace(day=1)
        current_month_str = first_of_month.strftime("%Y-%m")
        
        monthly_expenses = float(
            Transaction.objects.filter(
                user=request.user, txn_type="expense", date__gte=first_of_month
            ).aggregate(total=Sum("amount"))["total"] or 0
        )

        # Check for excess income
        excess_income_obj = ExcessIncome.objects.filter(user=request.user, month=current_month_str).first()
        excess_income = float(excess_income_obj.amount) if excess_income_obj else 0

        # Same fallback logic as the savings view (including excess income)
        if salary > 0:
            available = (salary + excess_income) - monthly_expenses
        else:
            monthly_income = float(
                Transaction.objects.filter(
                    user=request.user, txn_type="income", date__gte=first_of_month
                ).aggregate(total=Sum("amount"))["total"] or 0
            )
            available = monthly_income - monthly_expenses
        allocatable = max(available, 0)

        goals = list(SavingsGoal.objects.filter(user=request.user))

        for g in goals:
            if g.last_allocated_month != current_month_str:
                target = float(g.target_amount)
                base_saved = float(g.saved_amount)
                if base_saved > 0:
                    g.saved_amount = Decimal("0")
                    g.current_month_auto_allocation = Decimal(str(round(min(base_saved, target), 2)))
                    g.save(update_fields=["saved_amount", "current_month_auto_allocation"])
                else:
                    g.current_month_auto_allocation = Decimal("0")
                    g.save(update_fields=["current_month_auto_allocation"])

        available_after_prior = allocatable
        for g in goals:
            available_after_prior -= float(g.current_month_auto_allocation)
        available_after_prior = max(available_after_prior, 0)

        allocations = _calculate_goal_allocations(available_after_prior, goals)

        for g in goals:
            allocated = allocations.get(g.id, 0)
            if allocated > 0:
                effective_saved = float(g.saved_amount) + float(g.current_month_auto_allocation)
                new_effective = min(effective_saved + allocated, float(g.target_amount))
                new_auto = new_effective - float(g.saved_amount)
                g.current_month_auto_allocation = Decimal(str(round(new_auto, 2)))
                g.last_allocated_month = current_month_str
                g.save(update_fields=["current_month_auto_allocation", "last_allocated_month"])

        allocation_data = []
        for goal in goals:
            effective_saved = float(goal.saved_amount) + float(goal.current_month_auto_allocation)
            effective_saved = min(effective_saved, float(goal.saved_amount) + allocatable)
            allocation_data.append(
                {
                    "id": goal.id,
                    "name": goal.name,
                    "target_amount": float(goal.target_amount),
                    "saved_amount": effective_saved,
                    "remaining": max(float(goal.target_amount) - effective_saved, 0),
                    "priority": goal.priority,
                    "allocation_percentage": goal.allocation_percentage,
                    "allocated_this_month": allocations.get(goal.id, 0),
                    "is_complete": effective_saved >= float(goal.target_amount),
                }
            )

        return JsonResponse(
            {
                "success": True,
                "salary": salary,
                "excess_income": excess_income,
                "monthly_expenses": monthly_expenses,
                "available": available,
                "total_goals": len([g for g in goals if g.is_active_goal]),
                "allocations": allocation_data,
            }
        )
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def _goal_json(goal: "SavingsGoal") -> dict:
    CIRC = 263.9
    pct = goal.progress_pct
    return {
        "id": goal.id,
        "name": goal.name,
        "target_amount": float(goal.target_amount),
        "saved_amount": float(goal.saved_amount),
        "remaining": float(goal.remaining),
        "progress_pct": pct,
        "is_complete": goal.is_complete,
        "is_active": goal.is_active,
        "priority": goal.priority,
        "allocation_percentage": goal.allocation_percentage,
        "ring_dash": round(pct / 100 * CIRC, 1),
        "ring_gap": round(CIRC - pct / 100 * CIRC, 1),
    }


def _calculate_goal_allocations(available_amount: float, goals: list) -> dict:
    """
    Calculate how much to allocate to each goal based on priority weights.
    High=3, Medium=2, Low=1 — proportional split of available funds.
    Each goal is capped at its remaining amount needed.

    Returns dict with goal_id -> allocation_amount
    """
    if not goals or available_amount <= 0:
        return {}

    from datetime import date
    current_month_str = date.today().replace(day=1).strftime("%Y-%m")

    def base_saved_amount(goal):
        saved_value = float(goal.saved_amount or 0)
        if goal.last_allocated_month == current_month_str:
            saved_value = max(
                saved_value - float(getattr(goal, "current_month_auto_allocation", 0) or 0),
                0,
            )
        return saved_value

    def base_remaining_amount(goal):
        return max(float(goal.target_amount) - base_saved_amount(goal), 0)

    active_goals = [g for g in goals if g.is_active and base_remaining_amount(g) > 0]

    if not active_goals:
        return {}

    priority_weights = {"high": 3, "medium": 2, "low": 1}

    sorted_goals = sorted(
        active_goals,
        key=lambda g: priority_weights.get(g.priority, 2),
        reverse=True,
    )

    total_weight = sum(priority_weights.get(g.priority, 2) for g in sorted_goals)
    if total_weight == 0:
        total_weight = len(sorted_goals)

    allocations = {}
    remaining_pool = available_amount

    for goal in sorted_goals:
        if remaining_pool <= 0:
            allocations[goal.id] = 0.0
            continue
        weight = priority_weights.get(goal.priority, 2)
        share = weight / total_weight
        allocation = available_amount * share
        remaining_needed = base_remaining_amount(goal)
        allocation = min(allocation, remaining_needed, remaining_pool)
        allocation = max(allocation, 0.0)
        allocations[goal.id] = round(allocation, 2)
        remaining_pool -= allocation

    if remaining_pool > 0.01:
        for goal in sorted_goals:
            if remaining_pool <= 0:
                break
            still_needed = base_remaining_amount(goal) - allocations.get(goal.id, 0)
            if still_needed > 0:
                extra = min(still_needed, remaining_pool)
                allocations[goal.id] = round(allocations.get(goal.id, 0) + extra, 2)
                remaining_pool -= extra

    return allocations


# ── Profile ───────────────────────────────────────────────
@login_required(login_url="/login/")
def profile_view(request: HttpRequest) -> HttpResponse:
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    success = False

    if request.method == "POST":
        first_name = request.POST.get("first_name", "").strip()
        last_name = request.POST.get("last_name", "").strip()
        email = request.POST.get("email", "").strip().lower()

        request.user.first_name = first_name
        request.user.last_name = last_name
        if email and email != request.user.email:
            if (
                not User.objects.filter(email=email)
                .exclude(pk=request.user.pk)
                .exists()
            ):
                request.user.email = email
                request.user.username = email
        request.user.save()

        if "avatar" in request.FILES:
            profile.avatar = request.FILES["avatar"]
        profile.save()
        success = True

    return render(
        request,
        "login/profile.html",
        {
            "user": request.user,
            "profile": profile,
            "active_nav": "profile",
            "success": success,
        },
    )


# ── Monthly Analysis ──────────────────────────────────────
@login_required(login_url="/login/")
def monthly(request: HttpRequest) -> HttpResponse:
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    monthly_analysis = _build_monthly_analysis(
        request.user,
        profile,
        _parse_month_param(request.GET.get("month")),
    )
    return render(
        request,
        "login/monthly.html",
        {
            "user": request.user,
            "profile": profile,
            "active_nav": "monthly",
            **monthly_analysis,
        },
    )


# ── API: Email Monthly Analysis ───────────────────────────
@login_required(login_url="/login/")
@require_POST
def api_email_monthly_analysis(request: HttpRequest) -> JsonResponse:
    recipient = (request.user.email or "").strip()
    if not recipient:
        return JsonResponse(
            {"error": "Add an email address in your profile before sending reports."},
            status=400,
        )

    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    analysis = _build_monthly_analysis(
        request.user,
        profile,
        _parse_month_param(request.GET.get("month")),
    )
    email_context = _build_monthly_analysis_email_context(
        request.user, profile, analysis
    )
    subject = f"SpendWise Monthly Analysis - {analysis['selected_month_label']}"
    html_body = render_to_string("login/emails/monthly_analysis.html", email_context)
    text_body = strip_tags(html_body)

    email = EmailMultiAlternatives(
        subject=subject,
        body=text_body,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[recipient],
    )
    email.attach_alternative(html_body, "text/html")
    try:
        email.send(fail_silently=False)
    except Exception:
        return JsonResponse(
            {"error": "Email could not be sent. Please check SMTP settings."},
            status=502,
        )

    return JsonResponse(
        {
            "success": True,
            "message": f"Monthly analysis sent to {recipient}.",
        }
    )


# ── API: Export Monthly Report ─────────────────────────────
import csv
import io
from django.http import HttpResponse
from django.db.models import Sum


@login_required(login_url="/login/")
def api_export_monthly_pdf(request: HttpRequest) -> HttpResponse:
    month_param = request.GET.get("month")
    month_date = _parse_month_param(month_param)

    # Get month boundaries
    start_date = month_date.replace(day=1)
    if month_date.month == 12:
        end_date = month_date.replace(
            year=month_date.year + 1, month=1, day=1
        ) - timedelta(days=1)
    else:
        end_date = month_date.replace(month=month_date.month + 1, day=1) - timedelta(
            days=1
        )

    # Get transactions for the month
    transactions = Transaction.objects.filter(
        user=request.user, date__gte=start_date, date__lte=end_date
    ).order_by("-date")

    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    analysis = _build_monthly_analysis(request.user, profile, month_date)
    total_expense = analysis["total_expense"]
    if profile.salary:
        total_income = Decimal(str(profile.salary)) + analysis["excess_income"]
    else:
        total_income = analysis["total_income"]
    total_saved = total_income - total_expense

    def money_html(value):
        return f"Rs. {value:,.2f}"

    # Simple printable HTML response for PDF export.
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <title>Monthly Report - {month_date.strftime("%B %Y")}</title>
        <style>
            body {{ font-family: Arial, sans-serif; padding: 40px; }}
            h1 {{ color: #333; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }}
            th {{ background: #f5f5f5; }}
            .summary {{ margin: 20px 0; padding: 20px; background: #f9f9f9; border-radius: 8px; }}
            .positive {{ color: green; }}
            .negative {{ color: red; }}
        </style>
    </head>
    <body>
        <h1>Monthly Analysis Report</h1>
        <p>Month: {month_date.strftime("%B %Y")}</p>
        <div class="summary">
            <p><strong>Total Income:</strong> <span class="positive">{money_html(total_income)}</span></p>
            <p><strong>Total Expenses:</strong> <span class="negative">{money_html(total_expense)}</span></p>
            <p><strong>Net Savings:</strong> <span class="{"positive" if total_saved >= 0 else "negative"}">{money_html(total_saved)}</span></p>
        </div>
        <h2>Transactions</h2>
        <table>
            <tr><th>Date</th><th>Type</th><th>Category</th><th>Description</th><th>Amount</th></tr>
            {"".join(f"<tr><td>{t.date}</td><td>{t.txn_type}</td><td>{t.category}</td><td>{t.title}</td><td class={'positive' if t.txn_type == 'income' else 'negative'}>{'+' if t.txn_type == 'income' else '-'}{money_html(t.amount)}</td></tr>" for t in transactions)}
        </table>
    </body>
    </html>
    """

    response = HttpResponse(html_content, content_type="text/html; charset=utf-8")
    response["Content-Disposition"] = (
        f"inline; filename=Monthly_Analysis_{month_date.strftime('%Y-%m')}.html"
    )
    return response


@login_required(login_url="/login/")
def api_export_monthly_xlsx(request: HttpRequest) -> HttpResponse:
    try:
        import openpyxl
    except ImportError:
        return HttpResponse("Excel export not available. Install openpyxl.", status=500)

    month_param = request.GET.get("month")
    month_date = _parse_month_param(month_param)

    # Get month boundaries
    start_date = month_date.replace(day=1)
    if month_date.month == 12:
        end_date = month_date.replace(
            year=month_date.year + 1, month=1, day=1
        ) - timedelta(days=1)
    else:
        end_date = month_date.replace(month=month_date.month + 1, day=1) - timedelta(
            days=1
        )

    # Get transactions for the month
    transactions = Transaction.objects.filter(
        user=request.user, date__gte=start_date, date__lte=end_date
    ).order_by("-date")

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Report"

    # Summary section
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    analysis = _build_monthly_analysis(request.user, profile, month_date)
    total_expense = analysis["total_expense"]
    if profile.salary:
        total_income = Decimal(str(profile.salary)) + analysis["excess_income"]
    else:
        total_income = analysis["total_income"]
    total_saved = total_income - total_expense

    ws["A1"] = f"Monthly Analysis Report - {month_date.strftime('%B %Y')}"
    ws["A1"].font = openpyxl.styles.Font(bold=True, size=14)
    ws.merge_cells("A1:E1")

    ws["A3"] = "Summary"
    ws["A3"].font = openpyxl.styles.Font(bold=True)

    ws["A4"] = "Total Income"
    ws["B4"] = float(total_income)

    ws["A5"] = "Total Expenses"
    ws["B5"] = float(total_expense)

    ws["A6"] = "Net Savings"
    ws["B6"] = float(total_income - total_expense)

    # Transactions
    ws["A8"] = "Date"
    ws["B8"] = "Type"
    ws["C8"] = "Category"
    ws["D8"] = "Description"
    ws["E8"] = "Amount"
    for cell in ["A8", "B8", "C8", "D8", "E8"]:
        ws[cell].font = openpyxl.styles.Font(bold=True)

    row = 9
    for t in transactions:
        ws[f"A{row}"] = t.date.strftime("%Y-%m-%d")
        ws[f"B{row}"] = t.txn_type
        ws[f"C{row}"] = t.category
        ws[f"D{row}"] = t.title
        ws[f"E{row}"] = float(t.amount)
        row += 1

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f"attachment; filename=Monthly_Analysis_{month_date.strftime('%Y-%m')}.xlsx"
    )
    return response


@login_required(login_url="/login/")
def api_export_monthly_csv(request: HttpRequest) -> HttpResponse:
    month_param = request.GET.get("month")
    month_date = _parse_month_param(month_param)

    start_date = month_date.replace(day=1)
    if month_date.month == 12:
        end_date = month_date.replace(
            year=month_date.year + 1, month=1, day=1
        ) - timedelta(days=1)
    else:
        end_date = month_date.replace(month=month_date.month + 1, day=1) - timedelta(
            days=1
        )

    transactions = Transaction.objects.filter(
        user=request.user, date__gte=start_date, date__lte=end_date
    ).order_by("-date")

    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    analysis = _build_monthly_analysis(request.user, profile, month_date)
    total_expense = analysis["total_expense"]
    if profile.salary:
        total_income = Decimal(str(profile.salary)) + analysis["excess_income"]
    else:
        total_income = analysis["total_income"]
    total_saved = total_income - total_expense

    buffer = io.StringIO()
    writer = csv.writer(buffer)

    writer.writerow([f"Monthly Analysis Report - {month_date.strftime('%B %Y')}"])
    writer.writerow([])
    writer.writerow(["Summary"])
    writer.writerow(["Total Income", float(total_income)])
    writer.writerow(["Total Expenses", float(total_expense)])
    writer.writerow(["Net Savings", float(total_saved)])
    writer.writerow([])
    writer.writerow(["Date", "Type", "Category", "Description", "Amount"])

    for t in transactions:
        sign = "+" if t.txn_type == "income" else "-"
        writer.writerow([
            t.date.strftime("%Y-%m-%d"),
            t.txn_type,
            t.category,
            t.title,
            f"{sign}{float(t.amount):,.2f}",
        ])

    response = HttpResponse(
        "\ufeff" + buffer.getvalue(),
        content_type="text/csv; charset=utf-8",
    )
    response["Content-Disposition"] = (
        f"attachment; filename=Monthly_Analysis_{month_date.strftime('%Y-%m')}.csv"
    )
    return response


# ── API: Add Transaction ──────────────────────────────────
@login_required(login_url="/login/")
@require_POST
def api_add_transaction(request: HttpRequest) -> JsonResponse:
    try:
        data = json.loads(request.body)
        title = data.get("title", "").strip() or "Untitled"
        amount = data.get("amount", 0)
        txn_type = data.get("txn_type", "expense").strip()
        category = data.get("category", "other").strip()
        txn_date_str = data.get("date", str(date.today()))
        try:
            txn_date = date.fromisoformat(txn_date_str)
        except (ValueError, TypeError):
            txn_date = date.today()

        if txn_type not in ("income", "expense"):
            txn_type = "expense"
        if not amount or float(amount) <= 0:
            amount = 0

        txn = Transaction.objects.create(
            user=request.user,
            title=title,
            amount=Decimal(str(amount)),
            txn_type=txn_type,
            category=category,
            date=txn_date,
            note=data.get("note", ""),
        )

        profile, _ = UserProfile.objects.get_or_create(user=request.user)
        stats = _dashboard_stats(request.user, profile)
        saving_message, saving_message_variant = _chatbot_savings_message(
            profile,
            stats,
            request.session.get("last_chatbot_saving_variant"),
        )
        request.session["last_chatbot_saving_variant"] = saving_message_variant

        return JsonResponse(
            {
                "success": True,
                "txn": _serialize_txn(txn),
                "available_expense_dates": _available_expense_dates(request.user),
                "saving_message": saving_message,
                **_dashboard_stats_payload(stats),
            }
        )
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# ── API: Delete Transaction ───────────────────────────────
@login_required(login_url="/login/")
@require_http_methods(["DELETE"])
def api_delete_transaction(request: HttpRequest, txn_id: int) -> JsonResponse:
    try:
        txn = Transaction.objects.get(id=txn_id, user=request.user)
        txn.delete()
        profile, _ = UserProfile.objects.get_or_create(user=request.user)
        stats = _dashboard_stats(request.user, profile)
        return JsonResponse(
            {
                "success": True,
                "available_expense_dates": _available_expense_dates(request.user),
                **_dashboard_stats_payload(stats),
            }
        )
    except Transaction.DoesNotExist:
        return JsonResponse({"error": "Transaction not found."}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# ── API: Set Salary ───────────────────────────────────────
@login_required(login_url="/login/")
@require_POST
def api_set_salary(request: HttpRequest) -> JsonResponse:
    try:
        data = json.loads(request.body)
        salary = data.get("salary")
        if salary is None or salary == "":
            return JsonResponse({"error": "Salary is required."}, status=400)
        salary = float(salary)
        if salary < 0:
            return JsonResponse(
                {"error": "Salary must be a positive number."}, status=400
            )

        profile, _ = UserProfile.objects.get_or_create(user=request.user)
        profile.salary = salary
        profile.save()
        stats = _dashboard_stats(request.user, profile)

        return JsonResponse(
            {
                "success": True,
                "salary": float(profile.salary),
                **_dashboard_stats_payload(stats),
            }
        )
    except (ValueError, TypeError):
        return JsonResponse({"error": "Invalid salary value."}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# ── API: Set/Update Excess Income ─────────────────────────
@login_required(login_url="/login/")
@require_http_methods(["POST", "PUT", "DELETE"])
def api_excess_income(request: HttpRequest) -> JsonResponse:
    try:
        from .models import ExcessIncome
        today = date.today()
        current_month_str = today.replace(day=1).strftime("%Y-%m")
        
        if request.method == "DELETE":
            # Delete excess income for current month
            ExcessIncome.objects.filter(user=request.user, month=current_month_str).delete()
            profile, _ = UserProfile.objects.get_or_create(user=request.user)
            stats = _dashboard_stats(request.user, profile)
            return JsonResponse(
                {
                    "success": True,
                    "excess_income": 0,
                    "has_excess_income": False,
                    **_dashboard_stats_payload(stats),
                }
            )
        
        # POST or PUT - create or update
        data = json.loads(request.body)
        amount = data.get("amount")
        note = data.get("note", "").strip()
        
        if amount is None or amount == "":
            return JsonResponse({"error": "Amount is required."}, status=400)
        amount = float(amount)
        if amount <= 0:
            return JsonResponse(
                {"error": "Amount must be greater than zero."}, status=400
            )

        # Create or update excess income for current month
        excess_income, created = ExcessIncome.objects.update_or_create(
            user=request.user,
            month=current_month_str,
            defaults={"amount": Decimal(str(amount)), "note": note}
        )
        
        profile, _ = UserProfile.objects.get_or_create(user=request.user)
        stats = _dashboard_stats(request.user, profile)

        return JsonResponse(
            {
                "success": True,
                "excess_income": float(excess_income.amount),
                "has_excess_income": True,
                "note": excess_income.note,
                **_dashboard_stats_payload(stats),
            }
        )
    except (ValueError, TypeError):
        return JsonResponse({"error": "Invalid amount value."}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# ── API: Set Target Savings ───────────────────────────────
@login_required(login_url="/login/")
@require_POST
def api_set_target_savings(request: HttpRequest) -> JsonResponse:
    try:
        data = json.loads(request.body)
        target = data.get("target_savings")
        priority = data.get("priority", "medium")

        # Validate priority
        valid_priorities = ["high", "medium", "low"]
        if priority not in valid_priorities:
            priority = "medium"

        if target is None or target == "":
            return JsonResponse({"error": "Target savings is required."}, status=400)
        target = float(target)
        if target <= 0:
            return JsonResponse(
                {"error": "Target must be greater than zero."}, status=400
            )

        profile, _ = UserProfile.objects.get_or_create(user=request.user)
        profile.target_savings = target
        profile.priority = priority
        profile.save()

        stats = _dashboard_stats(request.user, profile)

        return JsonResponse(
            {
                "success": True,
                "target": float(profile.target_savings),
                "priority": profile.priority,
                "savings_pct": stats["savings_pct"],
                "ring_dash": stats["ring_dash"],
                "ring_gap": stats["ring_gap"],
            }
        )
    except (ValueError, TypeError):
        return JsonResponse({"error": "Invalid value."}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required(login_url="/login/")
@require_POST
def api_motivation_message(request: HttpRequest) -> JsonResponse:
    try:
        data = json.loads(request.body)
        saved_amount = float(data.get("saved_amount", 0) or 0)
        expense_amount = float(data.get("expense_amount", 0) or 0)
    except (ValueError, TypeError, json.JSONDecodeError):
        return JsonResponse({"error": "Invalid payload."}, status=400)

    tone = _motivation_tone(saved_amount, expense_amount)
    badge = MOTIVATION_BADGES.get(tone, "Start")
    quote, used_ai = _generate_motivation_quote(saved_amount, expense_amount, tone)

    return JsonResponse(
        {
            "success": True,
            "quote": quote,
            "tone": tone,
            "badge": badge,
            "used_ai": used_ai,
        }
    )


# ── API: Update Transaction ───────────────────────────────
@login_required(login_url="/login/")
@require_http_methods(["PUT"])
def api_update_transaction(request: HttpRequest, txn_id: int) -> JsonResponse:
    try:
        txn = Transaction.objects.get(id=txn_id, user=request.user)
        data = json.loads(request.body)

        title = data.get("title", "").strip() or "Untitled"
        amount = data.get("amount", 0)
        txn_type = data.get("txn_type", "expense").strip()
        category = data.get("category", "other").strip()
        txn_date_str = data.get("date", str(date.today()))

        try:
            txn_date = date.fromisoformat(txn_date_str)
        except (ValueError, TypeError):
            txn_date = date.today()

        if txn_type not in ("income", "expense"):
            txn_type = "expense"
        if not amount or float(amount) <= 0:
            return JsonResponse({"error": "Amount must be greater than 0."}, status=400)

        # Update the transaction
        txn.title = title
        txn.amount = Decimal(str(amount))
        txn.txn_type = txn_type
        txn.category = category
        txn.date = txn_date
        txn.note = data.get("note", "")
        txn.save()

        profile, _ = UserProfile.objects.get_or_create(user=request.user)
        stats = _dashboard_stats(request.user, profile)

        return JsonResponse(
            {
                "success": True,
                "txn": _serialize_txn(txn),
                "available_expense_dates": _available_expense_dates(request.user),
                **_dashboard_stats_payload(stats),
            }
        )
    except Transaction.DoesNotExist:
        return JsonResponse({"error": "Transaction not found."}, status=404)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)
